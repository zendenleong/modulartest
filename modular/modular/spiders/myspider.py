import requests
from bs4 import BeautifulSoup
import pyodbc
import pandas as pd
import scrapy
from scrapy import FormRequest
from scrapy.utils.response import open_in_browser
from openpyxl import load_workbook


class MyspiderSpider(scrapy.Spider):
    name = "myspider"
    allowed_domains = ["www.ccilindia.com"]
    start_urls = ["https://www.ccilindia.com/FPI_ARCV.aspx"]

    #testing out puting in specific date in the webscraping spider
    def parse(self, response):
        data = {
            'drpArchival': '15-Sep-2023'
        }
        yield FormRequest.from_response(response, formdata=data, callback=self.step2)

    def step2(self,response):
        data = {
            'btnFPISWH': 'Export To Excel'
        }
        yield FormRequest.from_response(response, formdata=data, callback=self.step3)

    def step3(self, response):
        # Save the response content as an Excel file
        filename = 'exported_data.xlsx'
        with open(filename, 'wb') as f:
            f.write(response.body)
        self.log(f'Saved {filename}')

        # Load the Excel file and clean the data
        wb = load_workbook(filename)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    # Remove HTML tags using a simple regex (replace with an empty string)
                    cell.value = cell.value.replace("<.*?>", "")

        # Save the cleaned Excel file
        clean_filename = 'cleaned_data.xlsx'
        wb.save(clean_filename)

        # Read the cleaned Excel file using pandas
        df = pd.read_excel(clean_filename)

        # Now, you can work with the DataFrame 'df' containing the cleaned Excel data
        # For example, you can print the first few rows:
        print(df.head())



'''
    def step3(self, response):
        # Check if the response contains the Excel content
        if 'Content-Disposition' in response.headers:
            # Assuming the response headers contain information about the Excel file
            filename = response.headers['Content-Disposition'].decode('utf-8').split('=')[1]
            with open(filename, 'wb') as f:
                f.write(response.body)
            self.log(f'Saved {filename}')
        else:
            self.log("Excel file not found in the response.")

        # You can open the response in the browser for debugging purposes
        open_in_browser(response)
'''